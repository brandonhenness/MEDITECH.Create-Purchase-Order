import ctypes
import openpyxl
import os
import tkinter as tk
from tkinter import filedialog

# Constants
VK_RETURN = 0x0D
VK_BACK = 0x08
VK_F12 = 0x7B

print(os.getcwd())

def find_window(class_name, window_name):
    hwnd = ctypes.windll.user32.FindWindowW(class_name, window_name)
    return hwnd

# Find window by partial window name
def find_window_by_partial_name(class_name, window_name):
    hwnd = ctypes.windll.user32.FindWindowW(class_name, None)
    while hwnd:
        if window_name in get_window_text(hwnd):
            return hwnd
        hwnd = ctypes.windll.user32.FindWindowExW(None, hwnd, class_name, None)
    return None

# List all the top level windows
def list_top_level_windows():
    hwnd = ctypes.windll.user32.GetTopWindow(None)
    while hwnd:
        print(f'handle: {hwnd} Text: {get_window_text(hwnd)} Class: {get_window_class_name(hwnd)}')
        hwnd = ctypes.windll.user32.GetWindow(hwnd, 2)

# Return a list of child windows for a given parent window using EnumChildWindows
def list_child_windows(hwnd_parent):
    hwnd_child_list = []

    def foreach_window(hwnd, lParam):
        if ctypes.windll.user32.IsWindowVisible(hwnd):
            hwnd_child_list.append(hwnd)
        return True

    ctypes.windll.user32.EnumChildWindows(hwnd_parent, ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.c_int, ctypes.c_int)(foreach_window), 0)
    return hwnd_child_list

# Find a child window by partial window name
def find_child_window_by_name(hwnd_parent, window_name):
    hwnd_child_list = list_child_windows(hwnd_parent)
    for hwnd in hwnd_child_list:
        if window_name in get_window_text(hwnd):
            return hwnd
    return None

# Find a child window by class name
def find_child_window_by_class_name(hwnd_parent, class_name):
    hwnd_child_list = list_child_windows(hwnd_parent)
    for hwnd in hwnd_child_list:
        if class_name == get_window_class_name(hwnd):
            return hwnd
    return None

# Get the text of a window first by giving it focus and then reading the text then restoring the focus to the previous window
def get_window_text(hwnd):
    # Get the current foreground window
    hwnd_foreground = ctypes.windll.user32.GetForegroundWindow()

    # Set the window to the foreground
    ctypes.windll.user32.SetForegroundWindow(hwnd)

    # Get the text
    buffer = ctypes.create_unicode_buffer(1024)
    ctypes.windll.user32.GetWindowTextW(hwnd, buffer, 1024)

    # Restore the foreground window
    ctypes.windll.user32.SetForegroundWindow(hwnd_foreground)

    return buffer.value

def get_window_class_name(hwnd):
    buffer = ctypes.create_unicode_buffer(1024)
    ctypes.windll.user32.GetClassNameW(hwnd, buffer, 1024)
    return buffer.value

# Send string of text to unfocused window
def send_string(hwnd, text):
    if not text:
        return
    for c in text:
        ctypes.windll.user32.SendMessageW(hwnd, 0x0102, ord(c), 0)

# Send a key to an unfocused window
def send_key(hwnd, key):
    ctypes.windll.user32.SendMessageW(hwnd, 0x0102, key, 0)

def create_purchase_order(hwnd, purchase_order):
    send_string(hwnd, purchase_order.purchase_order_number)
    send_key(hwnd, VK_RETURN)
    if purchase_order.purchase_order_number != 'N' and not check_for_confirmation_dialog():
        purchase_order.edit_mode = True
        return 
    send_key(hwnd, VK_RETURN)
    send_key(hwnd, VK_RETURN)
    if purchase_order.purchase_order_type != 'PURCHASE':
        for _ in range(len('PURCHASE')):
            send_key(hwnd, VK_BACK)
        send_string(hwnd, purchase_order.purchase_order_type)
    send_key(hwnd, VK_RETURN)
    if purchase_order.order_type != 'REGULAR':
        for _ in range(len('REGULAR')):
            send_key(hwnd, VK_BACK)
        send_string(hwnd, purchase_order.order_type)
    send_key(hwnd, VK_RETURN)
    if purchase_order.delivery_date:
        send_string(hwnd, purchase_order.delivery_date)
    else:
        send_string(hwnd, 'T')
    send_key(hwnd, VK_RETURN)
    if purchase_order.order_date:
        for _ in range(len(8)):
            send_key(hwnd, VK_BACK)
        send_string(hwnd, purchase_order.order_date)
    send_key(hwnd, VK_RETURN)
    if purchase_order.confirmation_date:
        for _ in range(len(8)):
            send_key(hwnd, VK_BACK)
        send_string(hwnd, purchase_order.confirmation_date)
    send_key(hwnd, VK_RETURN)
    if purchase_order.ok_to_prepay:
        send_key(hwnd, VK_BACK)
        send_string(hwnd, 'Y')
    send_key(hwnd, VK_RETURN)
    if purchase_order.buyer:
        for _ in range(10):
            send_key(hwnd, VK_BACK)
        send_string(hwnd, purchase_order.buyer)
    send_key(hwnd, VK_RETURN)
    if purchase_order.block_auto_update:
        send_key(hwnd, VK_BACK)
        send_string(hwnd, 'Y')
    send_key(hwnd, VK_RETURN)
    if purchase_order.order_type == 'STANDING':
        if purchase_order.auto_receive:
            send_key(hwnd, VK_BACK)
            send_string(hwnd, 'Y')
        send_key(hwnd, VK_RETURN)
        if purchase_order.days_in_cycle:
            send_string(hwnd, purchase_order.days_in_cycle)
        send_key(hwnd, VK_RETURN)
        if purchase_order.number_of_cycles:
            send_string(hwnd, purchase_order.number_of_cycles)
        send_key(hwnd, VK_RETURN)
    send_string(hwnd, purchase_order.vendor)
    send_key(hwnd, VK_RETURN)
    if purchase_order.order_type == 'BLANKET':
        send_key(hwnd, VK_RETURN)
        send_key(hwnd, VK_RETURN)
        send_string(hwnd, purchase_order.expire_date)
        send_key(hwnd, VK_RETURN)
        send_string(hwnd, purchase_order.maximum_price_per_order)
        send_key(hwnd, VK_RETURN)
        send_string(hwnd, purchase_order.maximum_price_total)
        send_key(hwnd, VK_RETURN)
        if purchase_order.EDI != 'GHXEDI':
            for _ in range(len('GHXEDI')):
                send_key(hwnd, VK_BACK)
            send_string(hwnd, purchase_order.EDI)
        send_key(hwnd, VK_RETURN)
        send_string(hwnd, purchase_order.global_location_number)
        send_key(hwnd, VK_RETURN)
        send_string(hwnd, purchase_order.ship_via)
        send_key(hwnd, VK_RETURN)
        if purchase_order.JIT:
            send_key(hwnd, VK_BACK)
            send_string(hwnd, 'Y')
        send_key(hwnd, VK_RETURN)
        send_string(hwnd, purchase_order.send_to_vendor)
    send_key(hwnd, VK_F12)
    send_key(hwnd, VK_BACK)
    send_string(hwnd, 'F')
    send_key(hwnd, VK_RETURN)

def create_purchase_order_lines(hwnd, purchase_order, purchase_order_lines):
    send_string(hwnd, '2')
    send_key(hwnd, VK_RETURN)
    for purchase_order_line in purchase_order_lines:
        create_purchase_order_line(hwnd, purchase_order, purchase_order_line)

def create_purchase_order_line(hwnd, purchase_order, purchase_order_line):
    if purchase_order.edit_mode:
        send_string(hwnd, purchase_order_line.line_number)
    else:
        send_string(hwnd, 'N')
    send_key(hwnd, VK_RETURN)
    
    if purchase_order_line.item_number:
        send_string(hwnd, purchase_order_line.item_number)
        send_key(hwnd, VK_RETURN)
        send_key(hwnd, VK_RETURN)
    else:
        send_string(hwnd, 'N')
        send_key(hwnd, VK_RETURN)
        send_string(hwnd, purchase_order_line.common_name)
        send_key(hwnd, VK_RETURN)
        send_string(hwnd, purchase_order_line.category)
        send_key(hwnd, VK_RETURN)
        send_string(hwnd, purchase_order_line.tax_code)
        send_key(hwnd, VK_RETURN)
        send_string(hwnd, purchase_order_line.vendor_catalog_number)
        send_key(hwnd, VK_RETURN)
        send_string(hwnd, purchase_order_line.manufacturer)
        send_key(hwnd, VK_RETURN)
        send_string(hwnd, purchase_order_line.manufacturer_catalog_number)
        send_key(hwnd, VK_RETURN)
        send_string(hwnd, purchase_order_line.GTIN)
        send_key(hwnd, VK_RETURN)
        send_string(hwnd, purchase_order_line.description1)
        send_key(hwnd, VK_RETURN)
        send_string(hwnd, purchase_order_line.description2)
    send_key(hwnd, VK_RETURN)
    send_string(hwnd, purchase_order_line.additional_description)
    send_key(hwnd, VK_RETURN)
    if purchase_order_line.inventory:
        send_string(hwnd, purchase_order_line.inventory)
    else:
        send_key(hwnd, VK_RETURN)
        send_string(hwnd, purchase_order_line.department)
    send_key(hwnd, VK_RETURN)
    send_string(hwnd, purchase_order_line.deliver_to)
    send_key(hwnd, VK_RETURN)
    if purchase_order_line.department:
        send_string(hwnd, purchase_order_line.EOC)
        send_key(hwnd, VK_RETURN)
    if purchase_order_line.GL_account:
        for _ in range(11):
            send_key(hwnd, VK_BACK)
        send_string(hwnd, purchase_order_line.GL_account)
        send_key(hwnd, VK_RETURN)
    if 'MISC' in purchase_order_line.item_number or purchase_order_line.item_number == 'N':
        send_string(hwnd, purchase_order_line.packaging_string)
        send_key(hwnd, VK_RETURN)
    if purchase_order_line.unit_of_purchase:
        if not 'MISC' in purchase_order_line.item_number or not purchase_order_line.item_number == 'N':
            for _ in range(3):
                send_key(hwnd, VK_BACK)
        send_string(hwnd, purchase_order_line.unit_of_purchase)
    send_key(hwnd, VK_RETURN)
    send_string(hwnd, purchase_order_line.conversion_packaging)
    send_key(hwnd, VK_RETURN)
    if purchase_order_line.cost:
        if not 'MISC' in purchase_order_line.item_number or not purchase_order_line.item_number == 'N':
            for _ in range(10):
                send_key(hwnd, VK_BACK)
        send_string(hwnd, purchase_order_line.cost)
    send_key(hwnd, VK_RETURN)
    if purchase_order_line.price_confirmed == 'N':
        send_key(hwnd, VK_BACK)
        send_string(hwnd, 'N')
    send_key(hwnd, VK_RETURN)
    send_string(hwnd, purchase_order_line.quantity)
    send_key(hwnd, VK_RETURN)
    if purchase_order.order_type == 'STANDING':
        send_string(hwnd, purchase_order_line.quantity_per_order)
        send_key(hwnd, VK_RETURN)
        if purchase_order_line.total_quantity:
            for _ in range(10):
                send_key(hwnd, VK_BACK)
            send_string(hwnd, purchase_order_line.total_quantity)
        send_key(hwnd, VK_RETURN)
    
    # System Checks
    # This item has been discontinued. Order anyway?
    # Warning...Vendor for PO not primary vendor for item. Display vendor order?
    # Make this vendor the PRIMARY VENDOR for this item?
    # Vendor Not found for this Item. New?
    # Item is in stock (or on order) in at least 1 inventory - list them? N

def create_purchase_order_from_file(hwnd, file_name):
    purchase_order = PurchaseOrder()
    purchase_order_lines = []
    wb = openpyxl.load_workbook(file_name)
    sheet = wb['PurchaseOrderInfo']
    for row in sheet.iter_rows(min_row=1):
        if row[0].value == 'Purchase Order Number':
            if row[1].value:
                purchase_order.purchase_order_number = str(row[1].value)
        elif row[0].value == 'Purchase Order Type':
            if row[1].value:
                purchase_order.purchase_order_type = str(row[1].value)
        elif row[0].value == 'Return From Purchase Order Number':
            if row[1].value:
                purchase_order.return_from_purchase_order_number = str(row[1].value)
        elif row[0].value == 'Inventory':
            if row[1].value:
                purchase_order.inventory = str(row[1].value)
        elif row[0].value == 'Order Type':
            if row[1].value:
                purchase_order.order_type = str(row[1].value)
        elif row[0].value == 'Delivery Date':
            if row[1].value:
                purchase_order.delivery_date = str(row[1].value)
        elif row[0].value == 'Order Date':
            if row[1].value:
                purchase_order.order_date = str(row[1].value)
        elif row[0].value == 'Confirmation Date':
            if row[1].value:
                purchase_order.confirmation_date = str(row[1].value)
        elif row[0].value == 'OK to Prepay':
            if row[1].value:
                purchase_order.ok_to_prepay = str(row[1].value)
        elif row[0].value == 'Buyer':
            if row[1].value:
                purchase_order.buyer = str(row[1].value)
        elif row[0].value == 'Block Auto Update':
            if row[1].value:
                purchase_order.block_auto_update = str(row[1].value)
        elif row[0].value == 'Auto Receive':
            if row[1].value:
                purchase_order.auto_receive = str(row[1].value)
        elif row[0].value == 'Days In Cycle':
            if row[1].value:
                purchase_order.days_in_cycle = str(row[1].value)
        elif row[0].value == 'Number of Cycles':
            if row[1].value:
                purchase_order.number_of_cycles = str(row[1].value)
        elif row[0].value == 'Vendor':
            if row[1].value:
                purchase_order.vendor = str(row[1].value)
    print(purchase_order)
    sheet = wb['LineItems']
    for row in sheet.iter_rows(min_row=2):
        purchase_order_line = PurchaseOrderLine()
        if row[0].value:
            purchase_order_line.item_number = str(row[0].value)
        if row[1].value:
            purchase_order_line.common_name = str(row[1].value)
        if row[2].value:
            purchase_order_line.category = str(row[2].value)
        if row[3].value:    
            purchase_order_line.tax_code = str(row[3].value)
        if row[4].value:
            purchase_order_line.vendor_catalog_number = str(row[4].value)
        if row[5].value:
            purchase_order_line.manufacturer = str(row[5].value)
        if row[6].value:
            purchase_order_line.manufacturer_catalog_number = str(row[6].value)
        if row[7].value:
            purchase_order_line.GTIN = str(row[7].value)
        if row[8].value:
            purchase_order_line.description1 = str(row[8].value)
        if row[9].value:
            purchase_order_line.description2 = str(row[9].value)
        if row[10].value:
            purchase_order_line.additional_description = str(row[10].value)
        if row[11].value:
            purchase_order_line.inventory = str(row[11].value)
        if row[12].value:
            purchase_order_line.department = str(row[12].value)
        if row[13].value:
            purchase_order_line.deliver_to = str(row[13].value)
        if row[14].value:
            purchase_order_line.EOC = str(row[14].value)
        if row[15].value:
            purchase_order_line.GL = str(row[15].value)
        if row[16].value:
            purchase_order_line.packaging_string = str(row[16].value)
        if row[17].value:
            purchase_order_line.unit_of_purchase = str(row[17].value)
        if row[18].value:
            purchase_order_line.conversion_packaging = str(row[18].value)
        if row[19].value:
            purchase_order_line.cost = str(row[19].value)
        if row[20].value:
            purchase_order_line.quantity = str(row[20].value)
        if row[21].value:
            purchase_order_line.quantity_per_order = str(row[21].value)
        purchase_order_lines.append(purchase_order_line)
    wb.close()
    create_purchase_order(hwnd, purchase_order)
    create_purchase_order_lines(hwnd, purchase_order, purchase_order_lines)
    send_key(hwnd, VK_RETURN)
    send_key(hwnd, VK_RETURN)

class PurchaseOrder:
    def __init__(self):
        self.edit_mode = False
        self.purchase_order_number = 'N'
        self.purchase_order_type = 'PURCHASE'
        self.return_from_purchase_order_number = None
        self.inventory = None
        self.order_type = 'REGULAR'
        self.delivery_date = 'T'
        self.order_date = None
        self.confirmation_date = None
        self.ok_to_prepay = False
        self.buyer = None
        self.block_auto_update = False
        self.auto_receive = False
        self.days_in_cycle = '1'
        self.number_of_cycles = '1'
        self.vendor = None
        self.expire_date = None
        self.maximum_price_per_order = None
        self.maximum_price_total = None
        self.EDI_program = 'GHXEDI'
        self.global_location_number = None
        self.ship_via = None
        self.JIT = False
        self.send_new_edition = False
    
    def __str__(self):
        return f'Purchase Order Number: {self.purchase_order_number}'

class PurchaseOrderLine:
    def __init__(self):
        self.line_number = None
        self.item_number = None
        self.common_name = None
        self.category = None
        self.tax_code = None
        self.vendor_catalog_number = None
        self.manufacturer = None
        self.manufacturer_catalog_number = None
        self.GTIN = None
        self.description1 = None
        self.description2 = None
        self.additional_description = None
        self.inventory = None
        self.department = None
        self.deliver_to = None
        self.EOC = None
        self.GL_account = None
        self.packaging_string = None
        self.unit_of_purchase = None
        self.conversion_packaging = None
        self.cost = None
        self.price_confirmed = None
        self.quantity = None
        self.quantity_per_order = None
    
    def __str__(self):
        return f'Item Number: {self.item_number}'

def explore_files():
    files = filedialog.askopenfilenames(initialdir = os.getcwd(),title = "Select file",filetypes = (("xlsx files","*.xlsx"),("all files","*.*")))
    if files:
        return files
    else:
        return None

def check_for_confirmation_dialog():
    hwnd = find_window_by_partial_name(None, 'Confirmation')
    if hwnd:
        button_hwnd = find_child_window_by_name(hwnd, 'Yes')
        ctypes.windll.user32.SendMessageA(button_hwnd, 0x00F5, 0, 0)
        return True
    else:
        return False

def print_window_info(hwnd):
    print(f'handle: {hwnd} Text: {get_window_text(hwnd)} Class: {get_window_class_name(hwnd)}')

def main():
    hwnd = find_window_by_partial_name(None, 'MM.GRY')
    print_window_info(hwnd)

    child_hwnd = find_child_window_by_name(hwnd, 'MM Main Menu')
    print_window_info(child_hwnd)

    files = explore_files()
    if files:
        send_string(child_hwnd, '31')
        send_key(child_hwnd, VK_RETURN)
        for file in files:
            create_purchase_order_from_file(child_hwnd, file)


if __name__ == '__main__':
    main()